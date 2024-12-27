import json
import logging

from openpyxl import Workbook

import crunch_uml.schema as sch
from crunch_uml import db
from crunch_uml.renderers.renderer import Renderer, RendererRegistry

logger = logging.getLogger()


@RendererRegistry.register(
    "xlsx",
    descr="Renders Excel sheet where each tab corresponds to one of the tables in te datamodel.",
)
class XLSXRenderer(Renderer):
    def render(self, args, schema: sch.Schema):
        # sourcery skip: use-named-expression
        wb = Workbook()
        wb.remove(wb.active)  # type: ignore # Remove default sheet

        # Retrieve all models dynamically
        base = db.Base
        models = base.metadata.tables
        session = schema.get_session()

        # Laad de mapper (bijvoorbeeld als een JSON-string via args)
        column_mapper = json.loads(args.mapper) if args.mapper else {}

        for table_name, table in models.items():
            ws = wb.create_sheet(title=table_name)

            # Headers
            columns = ["id"] if "id" in table.columns else []
            columns.extend(sorted([c.name for c in table.columns if c.name != "id"]))

            # Pas de mapper toe op de kolomnamen
            mapped_columns = [column_mapper.get(col, col) for col in columns]

            # Schrijf de gemapte kolomnamen in de header
            for col_num, mapped_column in enumerate(mapped_columns, 1):
                ws.cell(row=1, column=col_num, value=mapped_column)

            # Model class associated with the table
            model = base.model_lookup_by_table_name(table_name)

            if model:  # Ensure there's an associated model class
                # Data
                for row_num, record in enumerate(
                    session.query(model).filter(model.schema_id == schema.schema_id).all(),
                    2,
                ):
                    for col_num, column in enumerate(columns, 1):
                        # Pas de mapper toe op de waarde als nodig
                        if column == 'domein_iv3' and (isinstance(model, db.Class) or isinstance(model, db.Enumeratie)):
                            package = session.query(db.Package).filter(db.Package.id == record.package_id).one_or_none()
                            value = package.domain_name if package is not None else None
                        else:
                            value = getattr(record, column)
                        ws.cell(row=row_num, column=col_num, value=value)

        wb.save(args.outputfile)
